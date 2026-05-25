from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .database_utils import get_staff_summary,get_student_summary,get_db_connection,get_department_detail,get_faculty_detail,get_level_detail,get_available_years,search_staff,search_students
import json
import mysql.connector
from .sheets_utils import get_service_statistics, get_formatted_statistics
from django.http import JsonResponse, HttpResponse
import datetime

# Excel export support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will be disabled.")

@login_required
def dashboard_home(request):
    """
    หน้าหลักสำหรับเลือกประเภท Dashboard
    """
    return render(request, 'dashboard/index.html')

@login_required
def staff_dashboard(request):
    """
    แสดง Dashboard ข้อมูลบุคลากร
    """
    # ดึงข้อมูลสรุปบุคลากร
    summary = get_staff_summary()
    
    if not summary:
        # กรณีเกิดข้อผิดพลาดในการดึงข้อมูล
        return render(request, 'dashboard/error.html', {'error_message': 'ไม่สามารถดึงข้อมูลบุคลากรได้'})
    
    # เตรียมข้อมูลทั้งหมด - ไม่จำกัดเฉพาะ 10 อันดับแรก
    all_departments = summary['department_distribution']
    
    # แยกหมวดหมู่หน่วยงาน
    major_departments = [dept for dept in all_departments if dept['count'] >= 50]  # หน่วยงานใหญ่
    medium_departments = [dept for dept in all_departments if 10 <= dept['count'] < 50]  # หน่วยงานกลาง
    small_departments = [dept for dept in all_departments if 1 <= dept['count'] < 10]  # หน่วยงานเล็ก
    
    # ข้อมูลสำหรับกราฟ - แสดงทั้งหมด
    department_labels = [dept['DEPARTMENTNAME'] or 'ไม่ระบุ' for dept in all_departments]
    department_data = [dept['count'] for dept in all_departments]
    
    staff_type_labels = [stype['STFTYPENAME'] or 'ไม่ระบุ' for stype in summary['staff_type_distribution']]
    staff_type_data = [stype['count'] for stype in summary['staff_type_distribution']]
    
    gender_labels = [gender['GENDERNAMETH'] or 'ไม่ระบุ' for gender in summary['gender_distribution']]
    gender_data = [gender['count'] for gender in summary['gender_distribution']]
    
    context = {
        'total_staff': summary['total_staff'],
        # ข้อมูลกราฟ - แสดงทั้งหมด
        'department_labels': json.dumps(department_labels),
        'department_data': json.dumps(department_data),
        'staff_type_labels': json.dumps(staff_type_labels),
        'staff_type_data': json.dumps(staff_type_data),
        'gender_labels': json.dumps(gender_labels),
        'gender_data': json.dumps(gender_data),
        
        # ข้อมูลตาราง - แบ่งหมวดหมู่
        'all_departments': all_departments,
        'major_departments': major_departments,
        'medium_departments': medium_departments,
        'small_departments': small_departments,
        'total_departments': len(all_departments),
    }
    
    return render(request, 'dashboard/staff_dashboard.html', context)

@login_required
def student_dashboard(request):
    """
    แสดง Dashboard ข้อมูลนักศึกษา
    รองรับการกรองตามปีการศึกษา
    """
    # ตรวจสอบว่าผู้ใช้มีสิทธิ์เข้าถึง
    if hasattr(request.user, 'is_academic_service') and request.user.is_academic_service:
        # รับ parameter ปีการศึกษาจาก URL
        year_filter = request.GET.get('year')
        if year_filter and year_filter.isdigit():
            year_filter = int(year_filter)
        else:
            year_filter = None
            
        # ดึงข้อมูลสรุปนักศึกษา (กรองตามปีถ้ามี)
        summary = get_student_summary(year_filter)
        
        # ดึงรายการปีที่มีข้อมูล
        available_years = get_available_years()
        
        if not summary:
            # กรณีเกิดข้อผิดพลาดในการดึงข้อมูล
            return render(request, 'dashboard/error.html', {'error_message': 'ไม่สามารถดึงข้อมูลนักศึกษาได้'})
        
        # เตรียมข้อมูลสำหรับแสดงผลในกราฟ
        faculty_labels = [fac['faculty_name'] or 'ไม่ระบุ' for fac in summary['faculty_distribution']]
        faculty_data = [fac['count'] for fac in summary['faculty_distribution']]
        
        # จัดกลุ่มข้อมูล program สำหรับกราฟ (รวม level และ faculty เดียวกัน)
        program_summary = {}
        for prog in summary['program_distribution']:
            prog_name = prog['program_name'] or 'ไม่ระบุ'
            if prog_name in program_summary:
                program_summary[prog_name] += prog['count']
            else:
                program_summary[prog_name] = prog['count']
        
        # เรียงตาม count จากมากไปน้อย
        sorted_programs = sorted(program_summary.items(), key=lambda x: x[1], reverse=True)
        program_labels = [item[0] for item in sorted_programs]
        program_data = [item[1] for item in sorted_programs]
        
        level_labels = [level['level_name'] or 'ไม่ระบุ' for level in summary['education_level_distribution']]
        level_data = [level['count'] for level in summary['education_level_distribution']]
        
        gender_labels = [gender['gender'] for gender in summary['gender_distribution']]
        gender_data = [gender['count'] for gender in summary['gender_distribution']]
        
        year_labels = [str(year['year']) for year in summary['year_distribution']]
        year_data = [year['count'] for year in summary['year_distribution']]
        
        # เตรียมข้อมูลทั้งหมด - ไม่จำกัดเฉพาะ 10 อันดับแรก
        all_faculties = summary['faculty_distribution']
        all_programs = summary['program_distribution']
        
        # แยกหมวดหมู่คณะ
        major_faculties = [fac for fac in all_faculties if fac['count'] >= 500]  # คณะใหญ่
        medium_faculties = [fac for fac in all_faculties if 100 <= fac['count'] < 500]  # คณะกลาง
        small_faculties = [fac for fac in all_faculties if fac['count'] < 100]  # คณะเล็ก
        
        # แยกหมวดหมู่สาขาวิชา (ใช้ข้อมูลที่มี faculty_name แล้ว)
        major_programs = [prog for prog in all_programs if prog['count'] >= 50]  # สาขาใหญ่
        medium_programs = [prog for prog in all_programs if 10 <= prog['count'] < 50]  # สาขากลาง
        small_programs = [prog for prog in all_programs if prog['count'] < 10]  # สาขาเล็ก
        
        context = {
            'total_students': summary['total_students'],
            # ข้อมูลกราฟ - แสดงทั้งหมด
            'faculty_labels': json.dumps(faculty_labels),
            'faculty_data': json.dumps(faculty_data),
            'program_labels': json.dumps(program_labels),
            'program_data': json.dumps(program_data),
            'level_labels': json.dumps(level_labels),
            'level_data': json.dumps(level_data),
            'gender_labels': json.dumps(gender_labels),
            'gender_data': json.dumps(gender_data),
            'year_labels': json.dumps(year_labels),
            'year_data': json.dumps(year_data),
            
            # ข้อมูลตาราง - แบ่งหมวดหมู่
            'all_faculties': all_faculties,
            'major_faculties': major_faculties,
            'medium_faculties': medium_faculties,
            'small_faculties': small_faculties,
            'total_faculties': len(all_faculties),
            
            'all_programs': all_programs,
            'major_programs': major_programs,
            'medium_programs': medium_programs,
            'small_programs': small_programs,
            'total_programs': len(all_programs),
            
            # ข้อมูลระดับการศึกษา - สำหรับตารางสรุป
            'level_distribution': summary['education_level_distribution'],
            
            # ข้อมูลเพศ - สำหรับตารางสรุป
            'gender_distribution': summary['gender_distribution'],
            
            # ข้อมูลระดับการศึกษาทั้งหมด - สำหรับ filter
            'all_levels': summary['education_level_distribution'],
            
            # ข้อมูลสำหรับ Advanced Filter System
            'filter_data': {
                'faculties': [{'label': fac['faculty_name'] or 'ไม่ระบุ', 'value': fac['faculty_name'] or 'ไม่ระบุ', 'count': fac['count']} for fac in all_faculties],
                'levels': [{'label': level['level_name'] or 'ไม่ระบุ', 'value': level['level_name'] or 'ไม่ระบุ', 'count': level['count']} for level in summary['education_level_distribution']],
                'years': [{'label': f"พ.ศ. {year['buddhist_year']}", 'value': year['buddhist_year']} for year in available_years],
            },
            
            # ข้อมูล Year Filter
            'available_years': available_years,
            'selected_year': year_filter,
            'year_filter_label': f'พ.ศ. {year_filter}' if year_filter else 'ภาพรวมทั้งหมด',
        }
        
        return render(request, 'dashboard/student_dashboard.html', context)
    else:
        return render(request, 'dashboard/unauthorized.html')
    
def test_database_connection(request):
    """
    ทดสอบการเชื่อมต่อฐานข้อมูลและแสดงผลลัพธ์
    """
    if not request.user.is_superuser:
        return redirect('dashboard:home')
        
    connection = get_db_connection()
    status = {}
    
    if connection:
        status['connected'] = True
        
        # ทดสอบตาราง staff_info
        try:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM staff_info")
            record = cursor.fetchone()
            status['staff_count'] = record[0]
            status['staff_table_exists'] = True
        except mysql.connector.Error:
            status['staff_table_exists'] = False
            
        # ทดสอบตาราง students_info
        try:
            cursor = connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM students_info")
            record = cursor.fetchone()
            status['student_count'] = record[0]
            status['student_table_exists'] = True
        except mysql.connector.Error:
            status['student_table_exists'] = False
            
        connection.close()
    else:
        status['connected'] = False
    
    return render(request, 'dashboard/db_test.html', {'status': status})

@login_required
def service_statistics_view(request):
    """
    แสดงหน้าสถิติบริการ
    """
    # ดึงข้อมูลจาก Google Sheet
    all_stats = get_service_statistics()
    
    # แยกข้อมูลปี 2567 และ 2568 เท่านั้น (ไม่รวมปี 2566)
    data_2567 = [item for item in all_stats if item['year'] == '2567']
    data_2568 = [item for item in all_stats if item['year'] == '2568']
    
    print(f"ข้อมูลปี 2567: {len(data_2567)} รายการ")
    print(f"ข้อมูลปี 2568: {len(data_2568)} รายการ")
    
    # เรียงข้อมูลตามเดือน
    thai_month_order = {
        'ม.ค.': 1, 'ก.พ.': 2, 'มี.ค.': 3, 'เม.ย.': 4, 'พ.ค.': 5, 'มิ.ย.': 6,
        'ก.ค.': 7, 'ส.ค.': 8, 'ก.ย.': 9, 'ต.ค.': 10, 'พ.ย.': 11, 'ธ.ค.': 12
    }
    
    data_2567.sort(key=lambda x: thai_month_order.get(x['month'], 0))
    data_2568.sort(key=lambda x: thai_month_order.get(x['month'], 0))
    
    # เตรียมข้อมูลสำหรับกราฟ
    months = ['ม.ค.', 'ก.พ.', 'มี.ค.', 'เม.ย.', 'พ.ค.', 'มิ.ย.', 
              'ก.ค.', 'ส.ค.', 'ก.ย.', 'ต.ค.', 'พ.ย.', 'ธ.ค.']
    
    # เตรียมข้อมูลสำหรับกราฟเปรียบเทียบรายเดือน
    chart_data = {
        'labels': months,
        'data_2567': [0] * 12,
        'data_2568': [0] * 12
    }
    
    # เติมข้อมูลปี 2567
    for item in data_2567:
        month_index = thai_month_order.get(item['month'], 0) - 1
        if 0 <= month_index < 12:
            chart_data['data_2567'][month_index] = item['visit_count']
    
    # เติมข้อมูลปี 2568
    for item in data_2568:
        month_index = thai_month_order.get(item['month'], 0) - 1
        if 0 <= month_index < 12:
            chart_data['data_2568'][month_index] = item['visit_count']
    
    # คำนวณผลรวมรายปี
    sum_2567 = sum(item['visit_count'] for item in data_2567)
    sum_2568 = sum(item['visit_count'] for item in data_2568)
    
    # สรุปข้อมูล
    diff_percent = 0
    if sum_2567 > 0:
        diff_percent = ((sum_2568 - sum_2567) / sum_2567) * 100
    
    summary = {
        'total_2567': sum_2567,
        'total_2568': sum_2568,
        'diff_percent': diff_percent
    }
    
    print(f"สรุปข้อมูล: {summary}")
    print(f"ข้อมูลกราฟ: {chart_data}")
    
    # ส่งเฉพาะข้อมูลปี 2567 และ 2568 ไปยังเทมเพลต
    context = {
        'data_2567': data_2567,
        'data_2568': data_2568,
        'chart_data': chart_data,
        'summary': summary
    }
    
    return render(request, 'dashboard/service_statistics.html', context)

def test_sheets_api(request):
    from .sheets_utils import test_sheets_connection
    
    result = test_sheets_connection()
    
    return JsonResponse({
        'success': result,
        'message': 'Connection successful' if result else 'Connection failed'
    })

def test_raw_sheets_data(request):
    """
    ทดสอบดึงข้อมูลดิบจาก Google Sheets
    """
    from .sheets_utils import get_raw_sheet_data
    
    data = get_raw_sheet_data()
    
    return JsonResponse({
        'success': data is not None,
        'message': 'Data retrieved successfully' if data else 'Failed to retrieve data',
        'data': data
    })

@login_required
def department_detail(request, department_name):
    """
    แสดงรายละเอียดของหน่วยงานเฉพาะ
    """
    # Decode URL-encoded department name  
    import urllib.parse
    department_name = urllib.parse.unquote(department_name)
    
    # ดึงข้อมูลรายละเอียดหน่วยงาน
    dept_info = get_department_detail(department_name)
    
    if not dept_info:
        return render(request, 'dashboard/error.html', {
            'error_message': f'ไม่สามารถดึงข้อมูลหน่วยงาน "{department_name}" ได้'
        })
    
    # เตรียมข้อมูลสำหรับกราฟ
    gender_labels = [gender['GENDERNAMETH'] or 'ไม่ระบุ' for gender in dept_info['gender_distribution']]
    gender_data = [gender['count'] for gender in dept_info['gender_distribution']]
    
    position_labels = [pos['POSNAMETH'] or 'ไม่ระบุ' for pos in dept_info['position_distribution']]
    position_data = [pos['count'] for pos in dept_info['position_distribution']]
    
    employment_type_labels = [emp['STFTYPENAME'] or 'ไม่ระบุ' for emp in dept_info['employment_type_distribution']]
    employment_type_data = [emp['count'] for emp in dept_info['employment_type_distribution']]
    
    context = {
        'department_name': department_name,
        'total_staff': dept_info['total_staff'],
        'gender_labels': json.dumps(gender_labels),
        'gender_data': json.dumps(gender_data),
        'position_labels': json.dumps(position_labels),
        'position_data': json.dumps(position_data),
        'employment_type_labels': json.dumps(employment_type_labels),
        'employment_type_data': json.dumps(employment_type_data),
        'gender_distribution': dept_info['gender_distribution'],
        'position_distribution': dept_info['position_distribution'],
        'employment_type_distribution': dept_info['employment_type_distribution'],
        'staff_list': dept_info['staff_list']
    }
    
    return render(request, 'dashboard/department_detail.html', context)

@login_required
def export_department_excel(request, department_name):
    """
    Export ข้อมูลหน่วยงานเฉพาะ เป็น Excel (.xlsx)
    รวมข้อมูล: สรุปภาพรวม, ตำแหน่งงาน, รายชื่อบุคลากร
    """
    # ตรวจสอบว่า openpyxl พร้อมใช้งาน
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("ไม่สามารถ Export ได้ - กรุณาติดตั้ง openpyxl: pip install openpyxl", status=500)

    # Decode URL-encoded department name
    import urllib.parse
    department_name = urllib.parse.unquote(department_name)

    # ดึงข้อมูลรายละเอียดหน่วยงาน
    dept_info = get_department_detail(department_name)

    if not dept_info:
        return HttpResponse(f"ไม่สามารถดึงข้อมูลหน่วยงาน \"{department_name}\" ได้", status=500)

    # สร้าง Excel Workbook
    wb = Workbook()
    wb.remove(wb.active)  # ลบ default worksheet

    # กำหนดสไตล์
    header_font = Font(name='Tahoma', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    title_font = Font(name='Tahoma', size=14, bold=True, color='059669')
    data_font = Font(name='Tahoma', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    current_date = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")

    # ========== Sheet 1: สรุปภาพรวม ==========
    ws_summary = wb.create_sheet("สรุปภาพรวม")

    # หัวข้อรายงาน
    ws_summary['A1'] = f"รายงานข้อมูลบุคลากร - {department_name}"
    ws_summary['A1'].font = Font(name='Tahoma', size=16, bold=True)
    ws_summary['A2'] = f"วันที่ออกรายงาน: {current_date}"
    ws_summary['A3'] = "ระบบ AIMS - Academic Information Management System"

    # ข้อมูลสรุป
    summary_data = [
        ["หัวข้อ", "จำนวน", "หน่วย"],
        ["บุคลากรทั้งหมด", dept_info['total_staff'], "คน"],
        ["จำนวนตำแหน่งงาน", len(dept_info['position_distribution']), "ตำแหน่ง"],
        ["จำนวนประเภทการจ้าง", len(dept_info['employment_type_distribution']), "ประเภท"],
    ]

    for row_num, row_data in enumerate(summary_data, start=5):
        for col_num, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.font = data_font
            if row_num == 5:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            if col_num == 2 and row_num > 5:
                cell.alignment = center_align

    # สรุปตามเพศ
    row_start = 10
    ws_summary.cell(row=row_start, column=1, value="สรุปตามเพศ").font = title_font
    row_start += 1

    gender_headers = ["เพศ", "จำนวน", "สัดส่วน (%)"]
    for col_num, header in enumerate(gender_headers, start=1):
        cell = ws_summary.cell(row=row_start, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    for gender in dept_info['gender_distribution']:
        row_start += 1
        percentage = round((gender['count'] / dept_info['total_staff']) * 100, 2) if dept_info['total_staff'] > 0 else 0
        ws_summary.cell(row=row_start, column=1, value=gender['GENDERNAMETH'] or 'ไม่ระบุ').border = border
        ws_summary.cell(row=row_start, column=2, value=gender['count']).border = border
        ws_summary.cell(row=row_start, column=2).alignment = center_align
        ws_summary.cell(row=row_start, column=3, value=f"{percentage}%").border = border
        ws_summary.cell(row=row_start, column=3).alignment = center_align

    # สรุปประเภทการจ้าง
    row_start += 2
    ws_summary.cell(row=row_start, column=1, value="สรุปประเภทการจ้าง").font = title_font
    row_start += 1

    emp_headers = ["ประเภทการจ้าง", "จำนวน", "สัดส่วน (%)"]
    for col_num, header in enumerate(emp_headers, start=1):
        cell = ws_summary.cell(row=row_start, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    for emp in dept_info['employment_type_distribution']:
        row_start += 1
        percentage = round((emp['count'] / dept_info['total_staff']) * 100, 2) if dept_info['total_staff'] > 0 else 0
        ws_summary.cell(row=row_start, column=1, value=emp['STFTYPENAME'] or 'ไม่ระบุ').border = border
        ws_summary.cell(row=row_start, column=2, value=emp['count']).border = border
        ws_summary.cell(row=row_start, column=2).alignment = center_align
        ws_summary.cell(row=row_start, column=3, value=f"{percentage}%").border = border
        ws_summary.cell(row=row_start, column=3).alignment = center_align

    # ปรับขนาดคอลัมน์
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15

    # ========== Sheet 2: ตำแหน่งงาน ==========
    ws_position = wb.create_sheet("ตำแหน่งงาน")

    # หัวข้อ
    ws_position['A1'] = f"รายละเอียดตำแหน่งงาน - {department_name}"
    ws_position['A1'].font = Font(name='Tahoma', size=14, bold=True)
    ws_position['A2'] = f"จำนวนตำแหน่งทั้งหมด: {len(dept_info['position_distribution'])} ตำแหน่ง"

    # Headers
    position_headers = ["อันดับ", "ตำแหน่งงาน", "จำนวน (คน)", "สัดส่วน (%)"]
    for col_num, header in enumerate(position_headers, start=1):
        cell = ws_position.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    # Data rows
    for row_num, pos in enumerate(dept_info['position_distribution'], start=5):
        rank = row_num - 4
        percentage = round((pos['count'] / dept_info['total_staff']) * 100, 2) if dept_info['total_staff'] > 0 else 0

        ws_position.cell(row=row_num, column=1, value=rank).border = border
        ws_position.cell(row=row_num, column=1).alignment = center_align
        ws_position.cell(row=row_num, column=2, value=pos['POSNAMETH'] or 'ไม่ระบุ').border = border
        ws_position.cell(row=row_num, column=3, value=pos['count']).border = border
        ws_position.cell(row=row_num, column=3).alignment = center_align
        ws_position.cell(row=row_num, column=4, value=f"{percentage}%").border = border
        ws_position.cell(row=row_num, column=4).alignment = center_align

    # ปรับขนาดคอลัมน์
    ws_position.column_dimensions['A'].width = 10
    ws_position.column_dimensions['B'].width = 45
    ws_position.column_dimensions['C'].width = 15
    ws_position.column_dimensions['D'].width = 15

    # ========== Sheet 3: รายชื่อบุคลากร ==========
    ws_staff = wb.create_sheet("รายชื่อบุคลากร")

    # หัวข้อ
    ws_staff['A1'] = f"รายชื่อบุคลากร - {department_name}"
    ws_staff['A1'].font = Font(name='Tahoma', size=14, bold=True)
    ws_staff['A2'] = f"จำนวนบุคลากรทั้งหมด: {dept_info['total_staff']} คน"

    # Headers
    staff_headers = ["ลำดับ", "รหัสบุคลากร", "ชื่อ-นามสกุล", "เพศ", "ตำแหน่ง", "ประเภทการจ้าง"]
    for col_num, header in enumerate(staff_headers, start=1):
        cell = ws_staff.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    # Data rows
    for row_num, staff in enumerate(dept_info['staff_list'], start=5):
        rank = row_num - 4
        full_name = f"{staff.get('PREFIXFULLNAME', '')} {staff.get('STAFFNAME', '')} {staff.get('STAFFSURNAME', '')}".strip()

        ws_staff.cell(row=row_num, column=1, value=rank).border = border
        ws_staff.cell(row=row_num, column=1).alignment = center_align

        ws_staff.cell(row=row_num, column=2, value=staff.get('STAFFID', '')).border = border
        ws_staff.cell(row=row_num, column=2).alignment = center_align

        ws_staff.cell(row=row_num, column=3, value=full_name).border = border

        ws_staff.cell(row=row_num, column=4, value=staff.get('GENDERNAMETH', 'ไม่ระบุ')).border = border
        ws_staff.cell(row=row_num, column=4).alignment = center_align

        ws_staff.cell(row=row_num, column=5, value=staff.get('POSNAMETH', 'ไม่ระบุ')).border = border

        ws_staff.cell(row=row_num, column=6, value=staff.get('STFTYPENAME', 'ไม่ระบุ')).border = border

    # ปรับขนาดคอลัมน์
    ws_staff.column_dimensions['A'].width = 8
    ws_staff.column_dimensions['B'].width = 15
    ws_staff.column_dimensions['C'].width = 35
    ws_staff.column_dimensions['D'].width = 10
    ws_staff.column_dimensions['E'].width = 40
    ws_staff.column_dimensions['F'].width = 25

    # กำหนดชื่อไฟล์
    safe_dept_name = department_name.replace('/', '_').replace('\\', '_').replace(':', '_').replace(' ', '_')
    filename = f"AIMS_บุคลากร_{safe_dept_name}_{current_date.replace('/', '_').replace(' ', '_').replace(':', '')}.xlsx"

    # สร้าง HTTP Response สำหรับดาวน์โหลด
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@login_required
def faculty_detail(request, faculty_name):
    """
    แสดงรายละเอียดของคณะเฉพาะ
    รองรับการกรองตามปีการศึกษา
    """
    # Decode URL-encoded faculty name
    import urllib.parse
    faculty_name = urllib.parse.unquote(faculty_name)
    
    # รับ parameter ปีการศึกษาจาก URL
    year_filter = request.GET.get('year')
    if year_filter and year_filter.isdigit():
        year_filter = int(year_filter)
    else:
        year_filter = None
    
    # ดึงข้อมูลรายละเอียดคณะ (กรองตามปีถ้ามี)
    fac_info = get_faculty_detail(faculty_name, year_filter)
    
    if not fac_info:
        return render(request, 'dashboard/error.html', {
            'error_message': f'ไม่สามารถดึงข้อมูลคณะ "{faculty_name}" ได้'
        })
    
    # เตรียมข้อมูลสำหรับกราฟ
    gender_labels = [gender['gender'] or 'ไม่ระบุ' for gender in fac_info['gender_distribution']]
    gender_data = [gender['count'] for gender in fac_info['gender_distribution']]
    
    # จัดกลุ่มข้อมูล program สำหรับกราฟ (รวม level เดียวกัน)
    program_summary = {}
    for prog in fac_info['program_distribution']:
        prog_name = prog['program_name'] or 'ไม่ระบุ'
        if prog_name in program_summary:
            program_summary[prog_name] += prog['count']
        else:
            program_summary[prog_name] = prog['count']
    
    # เรียงตาม count จากมากไปน้อย
    sorted_programs = sorted(program_summary.items(), key=lambda x: x[1], reverse=True)
    program_labels = [item[0] for item in sorted_programs]
    program_data = [item[1] for item in sorted_programs]
    
    level_labels = [level['level_name'] or 'ไม่ระบุ' for level in fac_info['level_distribution']]
    level_data = [level['count'] for level in fac_info['level_distribution']]
    
    year_labels = [str(year['year']) for year in fac_info['year_distribution']]
    year_data = [year['count'] for year in fac_info['year_distribution']]
    
    context = {
        'faculty_name': faculty_name,
        'total_students': fac_info['total_students'],
        'gender_labels': json.dumps(gender_labels),
        'gender_data': json.dumps(gender_data),
        'program_labels': json.dumps(program_labels),
        'program_data': json.dumps(program_data),
        'level_labels': json.dumps(level_labels),
        'level_data': json.dumps(level_data),
        'year_labels': json.dumps(year_labels),
        'year_data': json.dumps(year_data),
        'gender_distribution': fac_info['gender_distribution'],
        'program_distribution': fac_info['program_distribution'],
        'level_distribution': fac_info['level_distribution'],
        'year_distribution': fac_info['year_distribution'],
        
        # Year Filter Information
        'selected_year': year_filter,
        'year_filter_label': f'พ.ศ. {year_filter}' if year_filter else 'ภาพรวมทั้งหมด',
    }
    
    return render(request, 'dashboard/faculty_detail.html', context)

@login_required
def level_detail(request, level_name):
    """
    แสดงรายละเอียดของระดับการศึกษาเฉพาะ
    รองรับการกรองตามปีการศึกษา
    """
    # Decode URL-encoded level name
    import urllib.parse
    level_name = urllib.parse.unquote(level_name)
    
    # รับ parameter ปีการศึกษาจาก URL
    year_filter = request.GET.get('year')
    if year_filter and year_filter.isdigit():
        year_filter = int(year_filter)
    else:
        year_filter = None
    
    # ดึงข้อมูลรายละเอียดระดับการศึกษา (กรองตามปีถ้ามี)
    level_info = get_level_detail(level_name, year_filter)
    
    if not level_info:
        return render(request, 'dashboard/error.html', {
            'error_message': f'ไม่สามารถดึงข้อมูลระดับการศึกษา "{level_name}" ได้'
        })
    
    # เตรียมข้อมูลสำหรับกราฟ
    gender_labels = [gender['gender'] or 'ไม่ระบุ' for gender in level_info['gender_distribution']]
    gender_data = [gender['count'] for gender in level_info['gender_distribution']]
    
    faculty_labels = [fac['faculty_name'] or 'ไม่ระบุ' for fac in level_info['faculty_distribution']]
    faculty_data = [fac['count'] for fac in level_info['faculty_distribution']]
    
    program_labels = [prog['program_name'] or 'ไม่ระบุ' for prog in level_info['program_distribution']]
    program_data = [prog['count'] for prog in level_info['program_distribution']]
    
    year_labels = [str(year['year']) for year in level_info['year_distribution']]
    year_data = [year['count'] for year in level_info['year_distribution']]
    
    context = {
        'level_name': level_name,
        'total_students': level_info['total_students'],
        'gender_labels': json.dumps(gender_labels),
        'gender_data': json.dumps(gender_data),
        'faculty_labels': json.dumps(faculty_labels),
        'faculty_data': json.dumps(faculty_data),
        'program_labels': json.dumps(program_labels),
        'program_data': json.dumps(program_data),
        'year_labels': json.dumps(year_labels),
        'year_data': json.dumps(year_data),
        'gender_distribution': level_info['gender_distribution'],
        'faculty_distribution': level_info['faculty_distribution'],
        'program_distribution': level_info['program_distribution'],
        'faculty_programs': level_info['faculty_programs'],
        'faculty_totals': level_info['faculty_totals'],
        'total_programs': level_info['total_programs'],
        'year_distribution': level_info['year_distribution'],
        
        # Year Filter Information
        'selected_year': year_filter,
        'year_filter_label': f'พ.ศ. {year_filter}' if year_filter else 'ภาพรวมทั้งหมด',
    }
    
    return render(request, 'dashboard/level_detail.html', context)

@login_required
def export_student_excel(request):
    """
    Export ข้อมูลนักศึกษาภาพรวม เป็น Excel
    รองรับการกรองตามปีการศึกษา
    """
    # ตรวจสอบสิทธิ์เข้าถึง
    if hasattr(request.user, 'is_academic_service') and request.user.is_academic_service:
        # รับ parameter ปีการศึกษาจาก URL
        year_filter = request.GET.get('year')
        if year_filter and year_filter.isdigit():
            year_filter = int(year_filter)
        else:
            year_filter = None
            
        # ดึงข้อมูลสรุปนักศึกษา (กรองตามปีถ้ามี)
        summary = get_student_summary(year_filter)
        
        if not summary:
            # กรณีเกิดข้อผิดพลาดในการดึงข้อมูล
            return HttpResponse("ไม่สามารถดึงข้อมูลได้", status=500)
        
        # สร้าง Excel Workbook
        wb = Workbook()
        
        # ลบ default worksheet
        wb.remove(wb.active)
        
        # กำหนดสไตล์
        header_font = Font(name='Tahoma', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        data_font = Font(name='Tahoma', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Sheet 1: สรุปภาพรวม
        ws_summary = wb.create_sheet("สรุปภาพรวม")
        current_date = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        year_text = f"พ.ศ. {year_filter}" if year_filter else "ทุกปีการศึกษา"
        
        ws_summary['A1'] = "รายงานสรุปข้อมูลนักศึกษา"
        ws_summary['A1'].font = Font(name='Tahoma', size=16, bold=True)
        ws_summary['A2'] = f"ข้อมูล: {year_text}"
        ws_summary['A3'] = f"วันที่ออกรายงาน: {current_date}"
        ws_summary['A4'] = "ระบบ AIMS - Academic Information Management System"
        
        # ข้อมูลสรุป
        summary_data = [
            ["หัวข้อ", "จำนวน", "หน่วย"],
            ["นักศึกษาทั้งหมด", summary['total_students'], "คน"],
            ["จำนวนคณะ", len(summary['faculty_distribution']), "คณะ"],
            ["จำนวนสาขาวิชา", len(summary['program_distribution']), "สาขา"],
            ["จำนวนระดับการศึกษา", len(summary['education_level_distribution']), "ระดับ"]
        ]
        
        for row_num, row_data in enumerate(summary_data, start=6):
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.font = data_font
                if row_num == 6:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                if col_num == 2 and row_num > 6:  # จำนวน column
                    cell.alignment = center_align
        
        # ปรับขนาดคอลัมน์
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 10
        
        # Sheet 2: ข้อมูลคณะ
        ws_faculty = wb.create_sheet("ข้อมูลคณะ")
        faculty_headers = ["ลำดับ", "ชื่อคณะ", "จำนวนนักศึกษา", "สัดส่วน (%)"]
        
        for col_num, header in enumerate(faculty_headers, start=1):
            cell = ws_faculty.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        for row_num, faculty in enumerate(summary['faculty_distribution'], start=2):
            percentage = round((faculty['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            row_data = [
                row_num - 1,
                faculty['faculty_name'] or 'ไม่ระบุ',
                faculty['count'],
                percentage
            ]
            
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_faculty.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.font = data_font
                if col_num in [1, 3, 4]:  # ลำดับ, จำนวน, สัดส่วน
                    cell.alignment = center_align
        
        # ปรับขนาดคอลัมน์
        ws_faculty.column_dimensions['A'].width = 8
        ws_faculty.column_dimensions['B'].width = 40
        ws_faculty.column_dimensions['C'].width = 18
        ws_faculty.column_dimensions['D'].width = 15
        
        # เพิ่มสรุปข้อมูลเพิ่มเติม
        current_row = len(summary['faculty_distribution']) + 3
        
        # 1. จำนวนนักศึกษาตามคณะทั้งหมด (Top 5)
        ws_faculty.cell(row=current_row, column=1, value="🏛️ Top 5 คณะที่มีนักศึกษามากที่สุด").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        
        top_5_faculties = summary['faculty_distribution'][:5]
        for i, faculty in enumerate(top_5_faculties, 1):
            ws_faculty.cell(row=current_row, column=1, value=f"{i}.")
            ws_faculty.cell(row=current_row, column=2, value=faculty['faculty_name'] or 'ไม่ระบุ')
            ws_faculty.cell(row=current_row, column=3, value=faculty['count'])
            percentage = round((faculty['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_faculty.cell(row=current_row, column=4, value=f"{percentage}%")
            current_row += 1
        
        current_row += 1
        
        # 2. สัดส่วนตามเพศ
        ws_faculty.cell(row=current_row, column=1, value="👥 สัดส่วนตามเพศ").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        
        for gender in summary['gender_distribution']:
            ws_faculty.cell(row=current_row, column=1, value="•")
            ws_faculty.cell(row=current_row, column=2, value=gender['gender'] or 'ไม่ระบุ')
            ws_faculty.cell(row=current_row, column=3, value=gender['count'])
            percentage = round((gender['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_faculty.cell(row=current_row, column=4, value=f"{percentage}%")
            current_row += 1
        
        current_row += 1
        
        # 3. ระดับการศึกษา
        ws_faculty.cell(row=current_row, column=1, value="🎓 ระดับการศึกษา").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        
        for level in summary['education_level_distribution']:
            ws_faculty.cell(row=current_row, column=1, value="•")
            ws_faculty.cell(row=current_row, column=2, value=level['level_name'] or 'ไม่ระบุ')
            ws_faculty.cell(row=current_row, column=3, value=level['count'])
            percentage = round((level['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_faculty.cell(row=current_row, column=4, value=f"{percentage}%")
            current_row += 1
        
        current_row += 1
        
        # 4. ปีที่เข้าศึกษา (Top 5)
        ws_faculty.cell(row=current_row, column=1, value="📅 Top 5 ปีที่เข้าศึกษา").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        
        top_5_years = summary['year_distribution'][:5]
        for i, year in enumerate(top_5_years, 1):
            ws_faculty.cell(row=current_row, column=1, value=f"{i}.")
            ws_faculty.cell(row=current_row, column=2, value=f"พ.ศ. {year['year']}" if year['year'] != 'ไม่ระบุ' else 'ไม่ระบุ')
            ws_faculty.cell(row=current_row, column=3, value=year['count'])
            percentage = round((year['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_faculty.cell(row=current_row, column=4, value=f"{percentage}%")
            current_row += 1
        
        # Sheet 3: ข้อมูลสาขาวิชา
        ws_program = wb.create_sheet("ข้อมูลสาขาวิชา")
        program_headers = ["ลำดับ", "คณะ", "ชื่อสาขาวิชา", "ระดับการศึกษา", "จำนวนนักศึกษา", "สัดส่วน (%)"]
        
        for col_num, header in enumerate(program_headers, start=1):
            cell = ws_program.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        for row_num, program in enumerate(summary['program_distribution'], start=2):
            percentage = round((program['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            row_data = [
                row_num - 1,
                program['faculty_name'] or 'ไม่ระบุ',
                program['program_name'] or 'ไม่ระบุ',
                program['level_name'] or 'ไม่ระบุ',
                program['count'],
                percentage
            ]
            
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_program.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.font = data_font
                if col_num in [1, 4, 5, 6]:  # ลำดับ, ระดับ, จำนวน, สัดส่วน
                    cell.alignment = center_align
                if col_num == 6 and isinstance(value, (int, float)):  # สัดส่วน column
                    cell.value = f"{value}%"
        
        # ปรับขนาดคอลัมน์
        ws_program.column_dimensions['A'].width = 8   # ลำดับ
        ws_program.column_dimensions['B'].width = 30  # คณะ
        ws_program.column_dimensions['C'].width = 45  # สาขาวิชา
        ws_program.column_dimensions['D'].width = 18  # ระดับการศึกษา
        ws_program.column_dimensions['E'].width = 15  # จำนวน
        ws_program.column_dimensions['F'].width = 12  # สัดส่วน
        
        # เพิ่มสรุปข้อมูลเพิ่มเติมในชีทสาขาวิชา
        current_row = len(summary['program_distribution']) + 3
        
        # 1. Top 5 คณะ
        ws_program.cell(row=current_row, column=1, value="🏛️ Top 5 คณะที่มีนักศึกษามากที่สุด").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        
        for i, faculty in enumerate(summary['faculty_distribution'][:5], 1):
            ws_program.cell(row=current_row, column=1, value=f"{i}.")
            ws_program.cell(row=current_row, column=2, value=faculty['faculty_name'] or 'ไม่ระบุ')
            ws_program.cell(row=current_row, column=3, value="")
            ws_program.cell(row=current_row, column=4, value="")
            ws_program.cell(row=current_row, column=5, value=faculty['count'])
            percentage = round((faculty['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_program.cell(row=current_row, column=6, value=f"{percentage}%")
            current_row += 1
        
        current_row += 1
        
        # 2. สัดส่วนตามเพศ
        ws_program.cell(row=current_row, column=1, value="👥 สัดส่วนตามเพศ").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        
        for gender in summary['gender_distribution']:
            ws_program.cell(row=current_row, column=1, value="•")
            ws_program.cell(row=current_row, column=2, value=gender['gender'] or 'ไม่ระบุ')
            ws_program.cell(row=current_row, column=3, value="")
            ws_program.cell(row=current_row, column=4, value="")
            ws_program.cell(row=current_row, column=5, value=gender['count'])
            percentage = round((gender['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_program.cell(row=current_row, column=6, value=f"{percentage}%")
            current_row += 1
        
        current_row += 1
        
        # 3. ระดับการศึกษา
        ws_program.cell(row=current_row, column=1, value="🎓 ระดับการศึกษา").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        
        for level in summary['education_level_distribution']:
            ws_program.cell(row=current_row, column=1, value="•")
            ws_program.cell(row=current_row, column=2, value="")
            ws_program.cell(row=current_row, column=3, value="")
            ws_program.cell(row=current_row, column=4, value=level['level_name'] or 'ไม่ระบุ')
            ws_program.cell(row=current_row, column=5, value=level['count'])
            percentage = round((level['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_program.cell(row=current_row, column=6, value=f"{percentage}%")
            current_row += 1
        
        current_row += 1
        
        # 4. Top 5 ปีที่เข้าศึกษา
        ws_program.cell(row=current_row, column=1, value="📅 Top 5 ปีที่เข้าศึกษา").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        
        for i, year in enumerate(summary['year_distribution'][:5], 1):
            ws_program.cell(row=current_row, column=1, value=f"{i}.")
            ws_program.cell(row=current_row, column=2, value="")
            ws_program.cell(row=current_row, column=3, value="")
            ws_program.cell(row=current_row, column=4, value=f"พ.ศ. {year['year']}" if year['year'] != 'ไม่ระบุ' else 'ไม่ระบุ')
            ws_program.cell(row=current_row, column=5, value=year['count'])
            percentage = round((year['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_program.cell(row=current_row, column=6, value=f"{percentage}%")
            current_row += 1
        
        # Sheet 4: ข้อมูลระดับการศึกษา
        ws_level = wb.create_sheet("ข้อมูลระดับการศึกษา")
        level_headers = ["ลำดับ", "ระดับการศึกษา", "จำนวนนักศึกษา", "สัดส่วน (%)"]
        
        for col_num, header in enumerate(level_headers, start=1):
            cell = ws_level.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        for row_num, level in enumerate(summary['education_level_distribution'], start=2):
            percentage = round((level['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            row_data = [
                row_num - 1,
                level['level_name'] or 'ไม่ระบุ',
                level['count'],
                percentage
            ]
            
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_level.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.font = data_font
                if col_num in [1, 3, 4]:  # ลำดับ, จำนวน, สัดส่วน
                    cell.alignment = center_align
        
        # ปรับขนาดคอลัมน์
        ws_level.column_dimensions['A'].width = 8
        ws_level.column_dimensions['B'].width = 25
        ws_level.column_dimensions['C'].width = 18
        ws_level.column_dimensions['D'].width = 15
        
        # เพิ่มสรุปข้อมูลเพิ่มเติมในชีทระดับการศึกษา
        current_row = len(summary['education_level_distribution']) + 3
        
        # 1. Top 5 คณะ
        ws_level.cell(row=current_row, column=1, value="🏛️ Top 5 คณะที่มีนักศึกษามากที่สุด").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for i, faculty in enumerate(summary['faculty_distribution'][:5], 1):
            ws_level.cell(row=current_row, column=1, value=f"{i}.").alignment = center_align
            ws_level.cell(row=current_row, column=2, value=faculty['faculty_name'] or 'ไม่ระบุ')
            ws_level.cell(row=current_row, column=3, value=faculty['count']).alignment = center_align
            percentage = round((faculty['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_level.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        current_row += 1
        
        # 2. สัดส่วนตามเพศ
        ws_level.cell(row=current_row, column=1, value="👥 สัดส่วนตามเพศ").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for gender in summary['gender_distribution']:
            ws_level.cell(row=current_row, column=1, value="•").alignment = center_align
            ws_level.cell(row=current_row, column=2, value=gender['gender'] or 'ไม่ระบุ')
            ws_level.cell(row=current_row, column=3, value=gender['count']).alignment = center_align
            percentage = round((gender['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_level.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        current_row += 1
        
        # 3. Top 5 สาขาวิชา
        ws_level.cell(row=current_row, column=1, value="📚 Top 5 สาขาวิชา").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for i, program in enumerate(summary['program_distribution'][:5], 1):
            ws_level.cell(row=current_row, column=1, value=f"{i}.").alignment = center_align
            ws_level.cell(row=current_row, column=2, value=program['program_name'] or 'ไม่ระบุ')
            ws_level.cell(row=current_row, column=3, value=program['count']).alignment = center_align
            percentage = round((program['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_level.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        current_row += 1
        
        # 4. Top 5 ปีที่เข้าศึกษา
        ws_level.cell(row=current_row, column=1, value="📅 Top 5 ปีที่เข้าศึกษา").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for i, year in enumerate(summary['year_distribution'][:5], 1):
            ws_level.cell(row=current_row, column=1, value=f"{i}.").alignment = center_align
            ws_level.cell(row=current_row, column=2, value=f"พ.ศ. {year['year']}" if year['year'] != 'ไม่ระบุ' else 'ไม่ระบุ')
            ws_level.cell(row=current_row, column=3, value=year['count']).alignment = center_align
            percentage = round((year['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_level.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        # Sheet 5: ข้อมูลเพศ
        ws_gender = wb.create_sheet("ข้อมูลเพศ")
        gender_headers = ["ลำดับ", "เพศ", "จำนวนนักศึกษา", "สัดส่วน (%)"]
        
        for col_num, header in enumerate(gender_headers, start=1):
            cell = ws_gender.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        for row_num, gender in enumerate(summary['gender_distribution'], start=2):
            percentage = round((gender['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            row_data = [
                row_num - 1,
                gender['gender'] or 'ไม่ระบุ',
                gender['count'],
                percentage
            ]
            
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_gender.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.font = data_font
                if col_num in [1, 3, 4]:  # ลำดับ, จำนวน, สัดส่วน
                    cell.alignment = center_align
        
        # ปรับขนาดคอลัมน์
        ws_gender.column_dimensions['A'].width = 8
        ws_gender.column_dimensions['B'].width = 15
        ws_gender.column_dimensions['C'].width = 18
        ws_gender.column_dimensions['D'].width = 15
        
        # เพิ่มสรุปข้อมูลเพิ่มเติมในชีทเพศ
        current_row = len(summary['gender_distribution']) + 3
        
        # 1. Top 5 คณะ
        ws_gender.cell(row=current_row, column=1, value="🏛️ Top 5 คณะที่มีนักศึกษามากที่สุด").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for i, faculty in enumerate(summary['faculty_distribution'][:5], 1):
            ws_gender.cell(row=current_row, column=1, value=f"{i}.").alignment = center_align
            ws_gender.cell(row=current_row, column=2, value=faculty['faculty_name'] or 'ไม่ระบุ')
            ws_gender.cell(row=current_row, column=3, value=faculty['count']).alignment = center_align
            percentage = round((faculty['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_gender.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        current_row += 1
        
        # 2. Top 5 สาขาวิชา
        ws_gender.cell(row=current_row, column=1, value="📚 Top 5 สาขาวิชา").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for i, program in enumerate(summary['program_distribution'][:5], 1):
            ws_gender.cell(row=current_row, column=1, value=f"{i}.").alignment = center_align
            ws_gender.cell(row=current_row, column=2, value=program['program_name'] or 'ไม่ระบุ')
            ws_gender.cell(row=current_row, column=3, value=program['count']).alignment = center_align
            percentage = round((program['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_gender.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        current_row += 1
        
        # 3. ระดับการศึกษา
        ws_gender.cell(row=current_row, column=1, value="🎓 ระดับการศึกษา").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for level in summary['education_level_distribution']:
            ws_gender.cell(row=current_row, column=1, value="•").alignment = center_align
            ws_gender.cell(row=current_row, column=2, value=level['level_name'] or 'ไม่ระบุ')
            ws_gender.cell(row=current_row, column=3, value=level['count']).alignment = center_align
            percentage = round((level['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_gender.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        current_row += 1
        
        # 4. Top 5 ปีที่เข้าศึกษา
        ws_gender.cell(row=current_row, column=1, value="📅 Top 5 ปีที่เข้าศึกษา").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for i, year in enumerate(summary['year_distribution'][:5], 1):
            ws_gender.cell(row=current_row, column=1, value=f"{i}.").alignment = center_align
            ws_gender.cell(row=current_row, column=2, value=f"พ.ศ. {year['year']}" if year['year'] != 'ไม่ระบุ' else 'ไม่ระบุ')
            ws_gender.cell(row=current_row, column=3, value=year['count']).alignment = center_align
            percentage = round((year['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_gender.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        # Sheet 6: ข้อมูลปีเข้าศึกษา
        ws_year = wb.create_sheet("ข้อมูลปีเข้าศึกษา")
        year_headers = ["ลำดับ", "ปี พ.ศ.", "จำนวนนักศึกษา", "สัดส่วน (%)"]
        
        for col_num, header in enumerate(year_headers, start=1):
            cell = ws_year.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        for row_num, year in enumerate(summary['year_distribution'], start=2):
            percentage = round((year['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            row_data = [
                row_num - 1,
                year['year'] if year['year'] != 'ไม่ระบุ' else 'ไม่ระบุ',
                year['count'],
                percentage
            ]
            
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_year.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.font = data_font
                if col_num in [1, 2, 3, 4]:  # ทุกคอลัมน์ center
                    cell.alignment = center_align
        
        # ปรับขนาดคอลัมน์
        ws_year.column_dimensions['A'].width = 8
        ws_year.column_dimensions['B'].width = 12
        ws_year.column_dimensions['C'].width = 18
        ws_year.column_dimensions['D'].width = 15
        
        # เพิ่มสรุปข้อมูลเพิ่มเติมในชีทปีเข้าศึกษา
        current_row = len(summary['year_distribution']) + 3
        
        # 1. Top 5 คณะ
        ws_year.cell(row=current_row, column=1, value="🏛️ Top 5 คณะที่มีนักศึกษามากที่สุด").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for i, faculty in enumerate(summary['faculty_distribution'][:5], 1):
            ws_year.cell(row=current_row, column=1, value=f"{i}.").alignment = center_align
            ws_year.cell(row=current_row, column=2, value=faculty['faculty_name'] or 'ไม่ระบุ')
            ws_year.cell(row=current_row, column=3, value=faculty['count']).alignment = center_align
            percentage = round((faculty['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_year.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        current_row += 1
        
        # 2. สัดส่วนตามเพศ
        ws_year.cell(row=current_row, column=1, value="👥 สัดส่วนตามเพศ").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for gender in summary['gender_distribution']:
            ws_year.cell(row=current_row, column=1, value="•").alignment = center_align
            ws_year.cell(row=current_row, column=2, value=gender['gender'] or 'ไม่ระบุ')
            ws_year.cell(row=current_row, column=3, value=gender['count']).alignment = center_align
            percentage = round((gender['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_year.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        current_row += 1
        
        # 3. ระดับการศึกษา
        ws_year.cell(row=current_row, column=1, value="🎓 ระดับการศึกษา").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for level in summary['education_level_distribution']:
            ws_year.cell(row=current_row, column=1, value="•").alignment = center_align
            ws_year.cell(row=current_row, column=2, value=level['level_name'] or 'ไม่ระบุ')
            ws_year.cell(row=current_row, column=3, value=level['count']).alignment = center_align
            percentage = round((level['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_year.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        current_row += 1
        
        # 4. Top 5 สาขาวิชา
        ws_year.cell(row=current_row, column=1, value="📚 Top 5 สาขาวิชา").font = Font(name='Tahoma', size=12, bold=True)
        current_row += 1
        for i, program in enumerate(summary['program_distribution'][:5], 1):
            ws_year.cell(row=current_row, column=1, value=f"{i}.").alignment = center_align
            ws_year.cell(row=current_row, column=2, value=program['program_name'] or 'ไม่ระบุ')
            ws_year.cell(row=current_row, column=3, value=program['count']).alignment = center_align
            percentage = round((program['count'] / summary['total_students']) * 100, 2) if summary['total_students'] > 0 else 0
            ws_year.cell(row=current_row, column=4, value=f"{percentage}%").alignment = center_align
            current_row += 1
        
        # กำหนดชื่อไฟล์
        if year_filter:
            filename = f"AIMS_นักศึกษา_พ.ศ.{year_filter}_{current_date.replace('/', '_').replace(' ', '_').replace(':', '')}.xlsx"
        else:
            filename = f"AIMS_นักศึกษา_ภาพรวมทั้งหมด_{current_date.replace('/', '_').replace(' ', '_').replace(':', '')}.xlsx"
        
        # สร้าง HTTP Response สำหรับดาวน์โหลด
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    else:
        return HttpResponse("ไม่มีสิทธิ์เข้าถึง", status=403)

@login_required
def export_faculty_excel(request, faculty_name):
    """
    Export ข้อมูลคณะเฉพาะ เป็น Excel
    รองรับการกรองตามปีการศึกษา
    """
    # Decode URL-encoded faculty name
    import urllib.parse
    faculty_name = urllib.parse.unquote(faculty_name)
    
    # ตรวจสอบสิทธิ์เข้าถึง
    if hasattr(request.user, 'is_academic_service') and request.user.is_academic_service:
        # รับ parameter ปีการศึกษาจาก URL
        year_filter = request.GET.get('year')
        if year_filter and year_filter.isdigit():
            year_filter = int(year_filter)
        else:
            year_filter = None
            
        # ดึงข้อมูลรายละเอียดคณะ (กรองตามปีถ้ามี)
        fac_info = get_faculty_detail(faculty_name, year_filter)
        
        if not fac_info:
            return HttpResponse("ไม่สามารถดึงข้อมูลได้", status=500)
        
        # Debug: ตรวจสอบข้อมูลที่ได้
        print(f"🔍 Faculty: {faculty_name}")
        print(f"🔍 Year Filter: {year_filter}")
        print(f"🔍 Total Students: {fac_info.get('total_students', 'N/A')}")
        print(f"🔍 Programs: {len(fac_info.get('program_distribution', []))}")
        print(f"🔍 Levels: {len(fac_info.get('level_distribution', []))}")
        
        # ตรวจสอบว่ามีข้อมูลหรือไม่
        if fac_info.get('total_students', 0) == 0:
            return HttpResponse("ไม่มีข้อมูลนักศึกษาในคณะนี้", status=400)
        
        # สร้าง Excel Workbook
        try:
            wb = Workbook()
        except Exception as e:
            print(f"❌ Excel Workbook Error: {e}")
            return HttpResponse(f"เกิดข้อผิดพลาดในการสร้าง Excel: {e}", status=500)
        
        # ลบ default worksheet
        wb.remove(wb.active)
        
        # กำหนดสไตล์
        header_font = Font(name='Tahoma', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        data_font = Font(name='Tahoma', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Sheet 1: สรุปภาพรวม
        ws_summary = wb.create_sheet("สรุปภาพรวม")
        current_date = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        year_text = f"พ.ศ. {year_filter}" if year_filter else "ทุกปีการศึกษา"
        
        ws_summary['A1'] = f"รายงานสรุปข้อมูลนักศึกษา - {faculty_name}"
        ws_summary['A1'].font = Font(name='Tahoma', size=16, bold=True)
        ws_summary['A2'] = f"ข้อมูล: {year_text}"
        ws_summary['A3'] = f"วันที่ออกรายงาน: {current_date}"
        ws_summary['A4'] = "ระบบ AIMS - Academic Information Management System"
        
        # ข้อมูลสรุป
        summary_data = [
            ["หัวข้อ", "จำนวน", "หน่วย"],
            ["นักศึกษาทั้งหมด", fac_info['total_students'], "คน"],
            ["จำนวนสาขาวิชา", len(fac_info['program_distribution']), "สาขา"],
            ["จำนวนระดับการศึกษา", len(fac_info['level_distribution']), "ระดับ"],
            ["จำนวนปีเข้าศึกษา", len(fac_info['year_distribution']), "ปี"]
        ]
        
        for row_num, row_data in enumerate(summary_data, start=6):
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.font = data_font
                if row_num == 6:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                if col_num == 2 and row_num > 6:  # จำนวน column
                    cell.alignment = center_align
        
        # ปรับขนาดคอลัมน์
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 10
        
        # Sheet 2: ข้อมูลสาขาวิชา
        ws_program = wb.create_sheet("ข้อมูลสาขาวิชา")
        program_headers = ["ลำดับ", "ชื่อสาขาวิชา", "ระดับการศึกษา", "จำนวนนักศึกษา", "สัดส่วน (%)"]
        
        for col_num, header in enumerate(program_headers, start=1):
            cell = ws_program.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        for row_num, program in enumerate(fac_info['program_distribution'], start=2):
            percentage = round((program['count'] / fac_info['total_students']) * 100, 2) if fac_info['total_students'] > 0 else 0
            row_data = [
                row_num - 1,
                program['program_name'] or 'ไม่ระบุ',
                program['level_name'] or 'ไม่ระบุ',
                program['count'],
                percentage
            ]
            
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_program.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.font = data_font
                if col_num in [1, 3, 4, 5]:  # ลำดับ, ระดับ, จำนวน, สัดส่วน
                    cell.alignment = center_align
                if col_num == 5 and isinstance(value, (int, float)):  # สัดส่วน column
                    cell.value = f"{value}%"
        
        # ปรับขนาดคอลัมน์
        ws_program.column_dimensions['A'].width = 8   # ลำดับ
        ws_program.column_dimensions['B'].width = 50  # สาขาวิชา
        ws_program.column_dimensions['C'].width = 18  # ระดับการศึกษา
        ws_program.column_dimensions['D'].width = 15  # จำนวน
        ws_program.column_dimensions['E'].width = 12  # สัดส่วน
        
        # Sheet 3: ข้อมูลระดับการศึกษา
        ws_level = wb.create_sheet("ข้อมูลระดับการศึกษา")
        level_headers = ["ลำดับ", "ระดับการศึกษา", "จำนวนนักศึกษา", "สัดส่วน (%)"]
        
        for col_num, header in enumerate(level_headers, start=1):
            cell = ws_level.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        for row_num, level in enumerate(fac_info['level_distribution'], start=2):
            percentage = round((level['count'] / fac_info['total_students']) * 100, 2) if fac_info['total_students'] > 0 else 0
            row_data = [
                row_num - 1,
                level['level_name'] or 'ไม่ระบุ',
                level['count'],
                percentage
            ]
            
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_level.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.font = data_font
                if col_num in [1, 3, 4]:  # ลำดับ, จำนวน, สัดส่วน
                    cell.alignment = center_align
                if col_num == 4 and isinstance(value, (int, float)):  # สัดส่วน column
                    cell.value = f"{value}%"
        
        # ปรับขนาดคอลัมน์
        ws_level.column_dimensions['A'].width = 8
        ws_level.column_dimensions['B'].width = 25
        ws_level.column_dimensions['C'].width = 18
        ws_level.column_dimensions['D'].width = 15
        
        # Sheet 4: ข้อมูลเพศ
        ws_gender = wb.create_sheet("ข้อมูลเพศ")
        gender_headers = ["ลำดับ", "เพศ", "จำนวนนักศึกษา", "สัดส่วน (%)"]
        
        for col_num, header in enumerate(gender_headers, start=1):
            cell = ws_gender.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        for row_num, gender in enumerate(fac_info['gender_distribution'], start=2):
            percentage = round((gender['count'] / fac_info['total_students']) * 100, 2) if fac_info['total_students'] > 0 else 0
            row_data = [
                row_num - 1,
                gender['gender'] or 'ไม่ระบุ',
                gender['count'],
                percentage
            ]
            
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_gender.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.font = data_font
                if col_num in [1, 3, 4]:  # ลำดับ, จำนวน, สัดส่วน
                    cell.alignment = center_align
                if col_num == 4 and isinstance(value, (int, float)):  # สัดส่วน column
                    cell.value = f"{value}%"
        
        # ปรับขนาดคอลัมน์
        ws_gender.column_dimensions['A'].width = 8
        ws_gender.column_dimensions['B'].width = 15
        ws_gender.column_dimensions['C'].width = 18
        ws_gender.column_dimensions['D'].width = 15
        
        # Sheet 5: ข้อมูลปีเข้าศึกษา
        ws_year = wb.create_sheet("ข้อมูลปีเข้าศึกษา")
        year_headers = ["ลำดับ", "ปี พ.ศ.", "จำนวนนักศึกษา", "สัดส่วน (%)"]
        
        for col_num, header in enumerate(year_headers, start=1):
            cell = ws_year.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        for row_num, year in enumerate(fac_info['year_distribution'], start=2):
            percentage = round((year['count'] / fac_info['total_students']) * 100, 2) if fac_info['total_students'] > 0 else 0
            row_data = [
                row_num - 1,
                year['year'] if year['year'] != 'ไม่ระบุ' else 'ไม่ระบุ',
                year['count'],
                percentage
            ]
            
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_year.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.font = data_font
                if col_num in [1, 2, 3, 4]:  # ทุกคอลัมน์ center
                    cell.alignment = center_align
                if col_num == 4 and isinstance(value, (int, float)):  # สัดส่วน column
                    cell.value = f"{value}%"
        
        # ปรับขนาดคอลัมน์
        ws_year.column_dimensions['A'].width = 8
        ws_year.column_dimensions['B'].width = 12
        ws_year.column_dimensions['C'].width = 18
        ws_year.column_dimensions['D'].width = 15
        
        # กำหนดชื่อไฟล์
        safe_faculty_name = faculty_name.replace('/', '_').replace('\\', '_').replace(':', '_')
        if year_filter:
            filename = f"AIMS_{safe_faculty_name}_พ.ศ.{year_filter}_{current_date.replace('/', '_').replace(' ', '_').replace(':', '')}.xlsx"
        else:
            filename = f"AIMS_{safe_faculty_name}_ภาพรวม_{current_date.replace('/', '_').replace(' ', '_').replace(':', '')}.xlsx"
        
        print(f"📁 Filename: {filename}")
        
        # สร้าง HTTP Response สำหรับดาวน์โหลด
        try:
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            print(f"✅ Excel file created successfully: {filename}")
            return response
        except Exception as e:
            print(f"❌ Excel Save Error: {e}")
            return HttpResponse(f"เกิดข้อผิดพลาดในการบันทึก Excel: {e}", status=500)
    else:
        return HttpResponse("ไม่มีสิทธิ์เข้าถึง", status=403)


@login_required
def export_staff_excel(request):
    """
    Export ข้อมูลบุคลากรภาพรวม เป็น Excel
    4 Sheets: สรุปภาพรวม, หน่วยงาน, ประเภทบุคลากร, เพศ
    """
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("ไม่สามารถ Export ได้ - กรุณาติดตั้ง openpyxl: pip install openpyxl", status=500)

    summary = get_staff_summary()
    if not summary:
        return HttpResponse("ไม่สามารถดึงข้อมูลบุคลากรได้", status=500)

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name='Tahoma', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    title_font  = Font(name='Tahoma', size=14, bold=True)
    data_font   = Font(name='Tahoma', size=11)
    bold_font   = Font(name='Tahoma', size=11, bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    current_date = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    total_staff  = summary['total_staff']
    all_depts    = summary['department_distribution']
    staff_types  = summary['staff_type_distribution']
    genders      = summary['gender_distribution']

    def write_headers(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = center

    def write_row(ws, row_num, values, font=None):
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col, value=v)
            c.font = font or data_font
            c.border = border
            c.alignment = center if isinstance(v, (int, float)) else left

    # ========== Sheet 1: สรุปภาพรวม ==========
    ws1 = wb.create_sheet("สรุปภาพรวม")
    ws1['A1'] = "รายงานสรุปข้อมูลบุคลากร"
    ws1['A1'].font = title_font
    ws1['A2'] = f"วันที่ออกรายงาน: {current_date}"
    ws1['A3'] = "ระบบ AIMS - Academic Information Management System"

    write_headers(ws1, ["หัวข้อ", "จำนวน", "หน่วย"], row=5)
    summary_rows = [
        ("บุคลากรทั้งหมด",       total_staff,                                              "คน"),
        ("จำนวนหน่วยงาน",        len(all_depts),                                           "หน่วยงาน"),
        ("หน่วยงานใหญ่ (50+)",   sum(1 for d in all_depts if d['count'] >= 50),            "หน่วยงาน"),
        ("หน่วยงานกลาง (10-49)", sum(1 for d in all_depts if 10 <= d['count'] < 50),       "หน่วยงาน"),
        ("หน่วยงานเล็ก (<10)",   sum(1 for d in all_depts if 1 <= d['count'] < 10),        "หน่วยงาน"),
        ("ประเภทบุคลากร",        len(staff_types),                                         "ประเภท"),
    ]
    for i, row in enumerate(summary_rows, 6):
        write_row(ws1, i, row)
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 12

    # ========== Sheet 2: หน่วยงาน ==========
    ws2 = wb.create_sheet("หน่วยงาน")
    ws2['A1'] = f"บุคลากรจำแนกตามหน่วยงาน ({len(all_depts)} หน่วยงาน)"
    ws2['A1'].font = title_font
    ws2['A2'] = f"วันที่: {current_date}"

    write_headers(ws2, ["ลำดับ", "ชื่อหน่วยงาน", "จำนวนบุคลากร", "สัดส่วน (%)"], row=4)
    for i, dept in enumerate(all_depts, 1):
        pct = round(dept['count'] / total_staff * 100, 2) if total_staff else 0
        write_row(ws2, i + 4, [i, dept['DEPARTMENTNAME'] or 'ไม่ระบุ', dept['count'], pct])
    total_row = len(all_depts) + 5
    write_row(ws2, total_row, ["", "รวมทั้งหมด", total_staff, 100.0], font=bold_font)
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 14

    # ========== Sheet 3: ประเภทบุคลากร ==========
    ws3 = wb.create_sheet("ประเภทบุคลากร")
    ws3['A1'] = "บุคลากรจำแนกตามประเภท"
    ws3['A1'].font = title_font
    ws3['A2'] = f"วันที่: {current_date}"

    write_headers(ws3, ["ลำดับ", "ประเภทบุคลากร", "จำนวน", "สัดส่วน (%)"], row=4)
    for i, stype in enumerate(staff_types, 1):
        pct = round(stype['count'] / total_staff * 100, 2) if total_staff else 0
        write_row(ws3, i + 4, [i, stype['STFTYPENAME'] or 'ไม่ระบุ', stype['count'], pct])
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 35
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 14

    # ========== Sheet 4: เพศ ==========
    ws4 = wb.create_sheet("เพศ")
    ws4['A1'] = "บุคลากรจำแนกตามเพศ"
    ws4['A1'].font = title_font
    ws4['A2'] = f"วันที่: {current_date}"

    write_headers(ws4, ["ลำดับ", "เพศ", "จำนวน", "สัดส่วน (%)"], row=4)
    for i, gender in enumerate(genders, 1):
        pct = round(gender['count'] / total_staff * 100, 2) if total_staff else 0
        write_row(ws4, i + 4, [i, gender['GENDERNAMETH'] or 'ไม่ระบุ', gender['count'], pct])
    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 14

    filename = f"AIMS_บุคลากร_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_level_excel(request, level_name):
    """
    Export ข้อมูลระดับการศึกษาเฉพาะ เป็น Excel
    3 Sheets: สรุปภาพรวม, สาขาวิชาจำแนกตามคณะ, เพศ
    """
    import urllib.parse
    level_name = urllib.parse.unquote(level_name)

    if not OPENPYXL_AVAILABLE:
        return HttpResponse("ไม่สามารถ Export ได้ - กรุณาติดตั้ง openpyxl", status=500)

    year_filter = request.GET.get('year')
    if year_filter and year_filter.isdigit():
        year_filter = int(year_filter)
    else:
        year_filter = None

    level_info = get_level_detail(level_name, year_filter)
    if not level_info:
        return HttpResponse("ไม่สามารถดึงข้อมูลได้", status=500)

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name='Tahoma', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    title_font  = Font(name='Tahoma', size=14, bold=True)
    data_font   = Font(name='Tahoma', size=11)
    bold_font   = Font(name='Tahoma', size=11, bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    current_date = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    year_text = f"พ.ศ. {year_filter}" if year_filter else "ทุกปีการศึกษา"
    total_students = level_info['total_students']

    def write_headers(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = center

    def write_row(ws, row_num, values, font=None):
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col, value=v)
            c.font = font or data_font
            c.border = border
            c.alignment = center if isinstance(v, (int, float)) else left

    # ========== Sheet 1: สรุปภาพรวม ==========
    ws1 = wb.create_sheet("สรุปภาพรวม")
    ws1['A1'] = f"รายงานสรุปข้อมูลนักศึกษา - {level_name}"
    ws1['A1'].font = title_font
    ws1['A2'] = f"ข้อมูล: {year_text}"
    ws1['A3'] = f"วันที่ออกรายงาน: {current_date}"
    ws1['A4'] = "ระบบ AIMS - Academic Information Management System"

    write_headers(ws1, ["หัวข้อ", "จำนวน", "หน่วย"], row=6)
    summary_rows = [
        ("นักศึกษาทั้งหมด",       total_students,                                  "คน"),
        ("จำนวนคณะ",               len(level_info['faculty_distribution']),          "คณะ"),
        ("จำนวนสาขาวิชา",          level_info['total_programs'],                     "สาขา"),
        ("ช่วงปีเข้าศึกษา",        len(level_info['year_distribution']),             "ปี"),
    ]
    for i, row in enumerate(summary_rows, 7):
        write_row(ws1, i, row)
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 10

    # ========== Sheet 2: สาขาวิชาจำแนกตามคณะ ==========
    ws2 = wb.create_sheet("สาขาวิชาจำแนกตามคณะ")
    ws2['A1'] = f"สาขาวิชาจำแนกตามคณะ - {level_name}"
    ws2['A1'].font = title_font
    ws2['A2'] = f"ข้อมูล: {year_text}  |  วันที่: {current_date}"

    write_headers(ws2, ["ลำดับ", "คณะ", "สาขาวิชา", "จำนวน (คน)", "สัดส่วน (%)"], row=4)

    row_num = 5
    seq = 1
    for faculty_name, programs in level_info['faculty_programs'].items():
        for prog in programs:
            pct = round(prog['count'] / total_students * 100, 2) if total_students else 0
            write_row(ws2, row_num, [seq, faculty_name, prog['program_name'] or 'ไม่ระบุ', prog['count'], pct])
            row_num += 1
            seq += 1
        # Faculty subtotal row
        ftotal = level_info['faculty_totals'].get(faculty_name, 0)
        fpct = round(ftotal / total_students * 100, 2) if total_students else 0
        write_row(ws2, row_num, ["", f"รวม {faculty_name}", "", ftotal, fpct], font=bold_font)
        row_num += 1

    # Grand total
    write_row(ws2, row_num, ["", "รวมทั้งหมด", "", total_students, 100.0], font=bold_font)
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 40
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 14

    # ========== Sheet 3: เพศ ==========
    ws3 = wb.create_sheet("เพศ")
    ws3['A1'] = f"นักศึกษาจำแนกตามเพศ - {level_name}"
    ws3['A1'].font = title_font
    ws3['A2'] = f"วันที่: {current_date}"

    write_headers(ws3, ["ลำดับ", "เพศ", "จำนวน (คน)", "สัดส่วน (%)"], row=4)
    for i, gender in enumerate(level_info['gender_distribution'], 1):
        pct = round(gender['count'] / total_students * 100, 2) if total_students else 0
        write_row(ws3, i + 4, [i, gender['gender'] or 'ไม่ระบุ', gender['count'], pct])
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 14

    safe_level = level_name.replace('/', '_').replace('\\', '_').replace(':', '_')
    if year_filter:
        filename = f"AIMS_{safe_level}_พ.ศ.{year_filter}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    else:
        filename = f"AIMS_{safe_level}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)


# ─────────────────────────────────────────────
# Sync Monitor Views
# ─────────────────────────────────────────────

@login_required
def sync_monitor_view(request):
    from .models import SyncLog

    last_staff = SyncLog.objects.filter(table_name='staff_info').first()
    last_students = SyncLog.objects.filter(table_name='students_info').first()

    staff_count = None
    students_count = None
    try:
        conn = get_db_connection()
        if conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM staff_info")
            staff_count = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM students_info")
            students_count = cur.fetchone()[0]
            cur.close()
            conn.close()
    except Exception:
        pass

    history = SyncLog.objects.all()[:30]

    context = {
        'last_staff': last_staff,
        'last_students': last_students,
        'staff_count': staff_count,
        'students_count': students_count,
        'history': history,
    }
    return render(request, 'dashboard/sync_monitor.html', context)


@login_required
def sync_trigger_api(request):
    import threading
    from .models import SyncLog
    from dashboard.management.commands.sync_staff import run_sync_staff
    from dashboard.management.commands.sync_students import run_sync_students

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    if request.user.user_role not in ('super_admin', 'staff_admin'):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    table = request.POST.get('table', '')
    if table not in ('staff_info', 'students_info', 'all'):
        return JsonResponse({'error': 'Invalid table'}, status=400)

    started_ids = []

    if table in ('staff_info', 'all'):
        log = SyncLog.objects.create(
            table_name='staff_info', status='running',
            triggered_by='manual', triggered_user=request.user,
        )
        started_ids.append({'table': 'staff_info', 'log_id': log.id})
        user = request.user
        staff_log = log
        t = threading.Thread(
            target=lambda: run_sync_staff(triggered_by='manual', triggered_user=user, existing_log=staff_log),
            daemon=True,
        )
        t.start()

    if table in ('students_info', 'all'):
        log = SyncLog.objects.create(
            table_name='students_info', status='running',
            triggered_by='manual', triggered_user=request.user,
        )
        started_ids.append({'table': 'students_info', 'log_id': log.id})
        user = request.user
        students_log = log
        t = threading.Thread(
            target=lambda: run_sync_students(triggered_by='manual', triggered_user=user, existing_log=students_log),
            daemon=True,
        )
        t.start()

    return JsonResponse({'status': 'started', 'jobs': started_ids})


@login_required
def sync_status_api(request, log_id):
    from .models import SyncLog
    try:
        log = SyncLog.objects.get(pk=log_id)
    except SyncLog.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)

    return JsonResponse({
        'status': log.status,
        'records_synced': log.records_synced,
        'records_before': log.records_before,
        'records_after': log.records_after,
        'duration_seconds': log.duration_seconds,
        'error_message': log.error_message,
        'finished_at': log.finished_at.isoformat() if log.finished_at else None,
    })
    return response

@login_required
def search_view(request):
    tab = request.GET.get('tab', 'staff')
    if tab not in ('staff', 'student'):
        tab = 'staff'
    query = request.GET.get('q', '').strip()
    results = []
    searched = False
    if query and len(query) >= 2:
        searched = True
        if tab == 'staff':
            results = search_staff(query)
        else:
            results = search_students(query)
        for row in results:
            raw = row.get('line_user_id')
            row['line_ids'] = raw.split(', ') if raw else []
    return render(request, 'dashboard/search.html', {
        'tab': tab,
        'query': query,
        'results': results,
        'searched': searched,
    })
